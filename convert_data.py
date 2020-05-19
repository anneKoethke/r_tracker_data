#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl as op
import json

print("--- loading data from excel ---")
# load excel (German formatting -> csv?)
path_to_wb = './data_input/input_data.xlsx' # muss noch umbenannt werden (sinnvoller FileName)
wb = op.load_workbook(path_to_wb)
sheet = wb['data']

data = []

id = 0
cat = ""
h3 = ""
state = ""
imgPath = ""
title = ""
abs_1 = ""
abs_2 = ""
link = ""
link_text = ""


print("--- processing data ---")

def get_cell_input(calling, r, c):
    if sheet.cell(row=r, column=c).value is None:
        # unwichtige Parameter für underscore
        if calling in ["Zwischenüberschrift", "Entwicklung (Abs. 2)", "Link"]:
            print("\tHinweis: kein Inhalt in", calling, "- Eintrag", r-1)
        else:
            print("FEHLER in Eintrag", r-1, ": kein Inhalt in", calling.upper(), "!")
        return ""
    else:
        return sheet.cell(row=r, column=c).value


def get_image_path(state, i):
    if state == "" or state is None:
        print("Fehler in 'Stand' von Eintrag", i-1, "(Excel-Zeile", i, ")")
        return ""
    elif state == "noch nicht begonnen":
        return "res/img/not.svg"
    elif state == "in Arbeit":
        return "res/img/half.svg"
    elif state == "erledigt":
        return "res/img/full.svg"
    elif state == "verfehlt":
        return "res/img/x.svg"
    else:
        print("Fehler in 'Stand' von Eintrag", i - 1, "(Excel-Zeile", i, ")")
        return ""


# process data: start_row, max_col (= Excel-Zeilenzahl + 1), rowwise
for i in range(2, 74, 1):
    id = i - 1
    cat = get_cell_input("Kategorie", i, 1)
    h3 = get_cell_input("Zwischenüberschrift", i, 2)
    state = get_cell_input("Stand", i, 3)
    imgPath = get_image_path(state, i)
    title = get_cell_input("Titel", i, 4)
    abs_1 = get_cell_input("Planung (Abs. 1)", i, 5)
    abs_2 = get_cell_input("Entwicklung (Abs. 2)", i, 6)
    link = get_cell_input("Link", i, 7)
    # prüfen, ob link vorhanden:
    if link == "":
        link_text = ""
    else:
        link_text = get_cell_input("Linkinfo", i, 8)

    data.append(
        {
            "id": id,
            "kategorie": cat,
            "h3": h3,
            "stand": state,
            "imagePath": imgPath,
            "titel": title,
            "versprechen": abs_1,
            "entwicklung": abs_2,
            "link": link,
            "linkinfo": link_text
        }
    )

# save as JSON or txt
print("--- writing to JSON ---")
with open('./data_output/output_data.json', 'w', encoding='utf-8') as json_file:
    json.dump(data, json_file, indent=2, ensure_ascii=False)
print("--- finished printing to JSON ---")